#!/usr/bin/env python
#
#The MIT License (MIT)
#
# Copyright (c) 2015 Bit9 + Carbon Black
#
# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the "Software"), to deal
# in the Software without restriction, including without limitation the rights
# to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
# copies of the Software, and to permit persons to whom the Software is
# furnished to do so, subject to the following conditions:
#
# The above copyright notice and this permission notice shall be included in all
# copies or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
# SOFTWARE.
#
# -----------------------------------------------------------------------------
#  This script is designed to output Process, Binary or Alert queries into a
#  multitabbed Excel SpreadSheet.
#
#  It takes a configuration file in the format of:
#  [Tab Name]
#  Type = binary|process|alert
#  Fields = <query output fields to display per column>
#  Query = <the Carbon Black query you want to run>
#
#  Thew output is completely customizable based on the values in the config
#  file.  This script was mainly designed for IR teams on proactive engagements 
#  to provide a leave behind to customers.
#
#  created 2016-03-02 by James Darby jdarby@carbonblack.com
#
# ------------------------------------------------------------------------------
#  TODO:
#
# ------------------------------------------------------------------------------

import sys
from cbapi.response import *
from cbapi.example_helpers import get_cb_response_object, build_cli_parser
from six.moves.configparser import ConfigParser
from six import PY3
from openpyxl import Workbook
import traceback


def convert_to_string(input):
    if not PY3:
        return unicode(input)
    else:
        return str(input)


def create_tab(section_name, myheader, wb):
    # prune the tab name to 31 for excel
    #
    tabname = section_name[:31]
    # create a unique tab name
    #
    wb.create_sheet(title=tabname)
    # grab the active worksheet
    #
    ws = wb.get_sheet_by_name(tabname)
    # print the header for each sheet
    #
    for col, this in enumerate(myheader):
        ws.cell(row=1, column=col + 1).value = this

    return ws


def main(argv):
    parser = build_cli_parser(description="Create an Excel report")
    parser.add_argument("-b", "--blank-tabs", action="store_true", default=True, dest="blanktabs",
                      help="Display blank tabs in Excel if the query returns no results.")
    parser.add_argument("-f", "--configfile", action="store", dest="configfile",
                      help="This is the configuration file", required=True)
    parser.add_argument("-o", "--outfile", action="store", default=None, dest="outfile",
                      help="This is the name of the spreadsheet that we'll create", required=True)

    args = parser.parse_args()
    cb = get_cb_response_object(args)

    config = ConfigParser()

    # build the Excel workbook
    #
    wb = Workbook()

    # read the config file in
    #
    config.readfp(open(args.configfile))

    # loop through the config file to pull each query
    #
    for section_name in config.sections():
        section_type = config.get(section_name, "type")
        if section_type == 'binary':
            query_cls = Binary
        elif section_type == 'process':
            query_cls = Process
        elif section_type == 'alert':
            query_cls = Alert
        else:
            print("Invalid query type {}, skipping.".format(section_type))
            continue

        # print the current query if in verbose mode
        #
        query = config.get(section_name, "query")
        if args.verbose:
            print('Working on query: {}'.format(query))

        created_tab = False
        myheader = config.get(section_name, "fields").split(",")
        myheader.insert(0,"Link")

        if args.blanktabs:
            ws = create_tab(section_name, myheader, wb)

        try:
            for row, result in enumerate(cb.select(query_cls).where(query)):
                if not args.blanktabs and not created_tab:
                    ws = create_tab(section_name, myheader, wb)
                    created_tab = True

                for col, header_title in enumerate(myheader):
                    if header_title == "Link":
                        if section_type == "alert":
                            link = result.proc.webui_link
                        else:
                            link = result.webui_link
                        ws.cell(row=row+2, column=col+1).hyperlink = link
                    else:
                        ws.cell(row=row+2, column=col+1).value = convert_to_string(result.get(header_title, "<UNKNOWN>"))
        except Exception as e:
            print("Encountered exception while processing query from section {0}: {1}".format(section_name, e))
            traceback.print_exc()
            print("Continuing...")

    # save the workbook
    #
    wb.remove_sheet(wb.get_sheet_by_name('Sheet'))
    wb.save(args.outfile)


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
